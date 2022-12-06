import logging
import time
import os
import warnings
import sys
from tqdm import tqdm

# Parameters used for searching
outputWorkbookName = 'lex-lme.xlsx'

lexHeader = {'name':'LEX (ALL DOMESTIC VOYAGES ONLY) REFUNDS/COMMISSIONS TO BE REIMBURSED BY CHECK/ WIRE (Payments to be processed through AS400)', 'tabTitle':'LEX'}
lmeHeader = {'name':'LME (All FOREIGN VOYAGES ONLY) REFUNDS/COMMISSIONS TO BE REIMBURSED BY CHECK/WIRE (Payments to be processed through AS400)', 'tabTitle':'LME'}

columns = ['BOOKING #', 'NAME', 'AMOUNT', 'COMMENTS', ' AMT PAID ', 'DATE PAID', 'CHK/CC', 'VOID DATE']

allSections = [lexHeader,lmeHeader]
sectionsToFind = allSections[:]
tabName = 'Sales'

# To log out EVERYTHING, change level from logging.INFO to logging.DEBUG
logginglevel = logging.INFO


# Set up logging and filter out the stupid UserWarning that header and footer can't be read
# To log out EVERYTHING, change level to logging.DEBUG
warnings.filterwarnings("ignore", category=UserWarning)
os.makedirs('logs', exist_ok=True)
starttime = time.strftime("%Y-%m-%d_%H.%M.%S")
logging.basicConfig(
    level=logginglevel,
    format='%(asctime)s %(levelname)s:  %(message)s',
    datefmt='%m/%d/%Y %I:%M:%S %p',
    encoding='utf-8',
    handlers=[
        logging.FileHandler(f'logs/{starttime}-looker.log'),
        logging.StreamHandler(sys.stdout)
    ])
logger = logging.getLogger('looker')

#Load in which workbooks to search
filesName = 'looker-files.txt'
workbooksToSearch = []
logger.debug(f'Workbooks to search:')
with open(filesName) as file:
    lines = file.readlines()
    for line in lines:
        line = line.strip('\n')
        line = line.strip('"')
        logger.debug(line)
        workbooksToSearch.append(line)

if len(workbooksToSearch) < 1:
    print('Error: At least one file must be present in looker-files.txt')
    logger.error('At least one file must be present in looker-files.txt')
    exit(1)

# This needs to be after opening up the other files cause of how I did it
from openpyxl import *

# Create our output workbook
outputWorkbook = Workbook()
# Remove any existing sheets/tabs
for sheet in outputWorkbook.sheetnames:
    outputWorkbook.remove(outputWorkbook[sheet])
# Create the tabs we'll need
for section in sectionsToFind:
    newsheet = outputWorkbook.create_sheet(section['tabTitle'])
    for i in range(len(columns)):
        newsheet.cell(row=1, column=i+1, value=columns[i])

# Search each workbook
for workbookName in tqdm(workbooksToSearch):
    logging.info(f'Processing workbook: {workbookName}')
    sectionsToFind = allSections[:]
    activeSection = None
    headerColumn = None
    workbook = load_workbook(filename=workbookName, read_only=True)
    searchSheet = workbook[tabName]
    for row in searchSheet.values:
        if activeSection is None:
            for section in sectionsToFind:
                if section['name'] in row:
                    #extract section
                    activeSection = section
                    logger.info(f'Found section: {activeSection["tabTitle"]}')
                    headerColumn = row.index(activeSection['name'])
                    sectionsToFind.remove(section)
                    break
        else:
            firstcell = row[headerColumn]
            if row[headerColumn] == columns[0]:
                logger.debug(f'Found a column header: {row[headerColumn]}')
                continue
            if row[headerColumn] is None and row[headerColumn+1] is None:
                logger.info(f'End of section: {activeSection["tabTitle"]}')
                activeSection = None
                headerColumn = None
                continue
            outputSheet = outputWorkbook[activeSection['tabTitle']]
            valuesToAdd = []
            for i in range (headerColumn,headerColumn+len(columns)):
                valuesToAdd.append(row[i])
            outputSheet.append(valuesToAdd)
            logger.debug(valuesToAdd)

os.makedirs('output', exist_ok=True)
outputWorkbookName = f'output\\{starttime}_{outputWorkbookName}'
outputWorkbook.save(outputWorkbookName)
print(f'Output stored in file: {os.path.join(os.getcwd(),outputWorkbookName)}')
